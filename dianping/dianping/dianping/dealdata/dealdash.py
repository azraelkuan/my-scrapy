# -*- coding: utf-8 -*-
import openpyxl
import re
import os
from dianping import pipelines

def deal_dashes():
    conn, cur = pipelines.connDB()
    sql = 'select * from shenzhen'
    cur.execute(sql)
    data = cur.fetchall()
    excel = openpyxl.Workbook()
    sheet = excel.create_sheet('shenzhen', 0)

    sheet.cell(row=1, column=1).value = 'name'
    sheet.cell(row=1, column=2).value = 'stars'
    sheet.cell(row=1, column=3).value = 'comments'
    sheet.cell(row=1, column=4).value = 'price'
    sheet.cell(row=1, column=5).value = 'taset'
    sheet.cell(row=1, column=6).value = 'environments'
    sheet.cell(row=1, column=7).value = 'service'
    sheet.cell(row=1, column=8).value = 'tel'
    sheet.cell(row=1, column=9).value = 'group purchase'
    sheet.cell(row=1, column=10).value = 'resevre'
    sheet.cell(row=1, column=11).value = 'take-out food'
    sheet.cell(row=1, column=12).value = 'promotion'
    sheet.cell(row=1, column=13).value = 'clubber'
    sheet.cell(row=1, column=14).value = 'opentime'
    sheet.cell(row=1, column=15).value = 'dish1'
    sheet.cell(row=1, column=16).value = 'dish1 price'
    sheet.cell(row=1, column=17).value = 'dish2'
    sheet.cell(row=1, column=18).value = 'dish2 price'
    sheet.cell(row=1, column=19).value = 'dish3'
    sheet.cell(row=1, column=20).value = 'dish3 price'
    sheet.cell(row=1, column=21).value = 'dish4'
    sheet.cell(row=1, column=22).value = 'dish4 price'
    sheet.cell(row=1, column=23).value = 'dish5'
    sheet.cell(row=1, column=24).value = 'dish5 price'
    sheet.cell(row=1, column=25).value = 'tag1'
    sheet.cell(row=1, column=26).value = 'tag2'
    sheet.cell(row=1, column=27).value = 'tag3'
    sheet.cell(row=1, column=28).value = 'tag4'
    sheet.cell(row=1, column=29).value = 'tag5'
    sheet.cell(row=1, column=30).value = 'url'
    sheet.cell(row=1, column=31).value = 'pro'
    sheet.cell(row=1, column=32).value = 'city'
    sheet.cell(row=1, column=33).value = 'district'
    sheet.cell(row=1, column=34).value = 'category_name'
    sheet.cell(row=1, column=35).value = 'category_sub_name'
    sheet.cell(row=1, column=36).value = 'address'

    for i in range(len(data)):
        j = i + 2
        print(i)
        sheet.cell(row=j, column=1).value = data[i][1]
        sheet.cell(row=j, column=2).value = data[i][2]
        sheet.cell(row=j, column=3).value = data[i][3]
        sheet.cell(row=j, column=4).value = data[i][4]
        sheet.cell(row=j, column=5).value = data[i][5]
        sheet.cell(row=j, column=6).value = data[i][6]
        sheet.cell(row=j, column=7).value = data[i][7]
        sheet.cell(row=j, column=8).value = data[i][8]
        sheet.cell(row=j, column=9).value = data[i][9]
        sheet.cell(row=j, column=10).value = data[i][10]
        sheet.cell(row=j, column=11).value = data[i][11]
        sheet.cell(row=j, column=12).value = data[i][12]
        sheet.cell(row=j, column=13).value = data[i][13]
        sheet.cell(row=j, column=14).value = data[i][14]

        dish_list = data[i][15]
        dish_list = dish_list.replace("no price", 'noprice').split(' ')
        print(dish_list)
        m = 0
        for each in dish_list:
            if each and m < 10:
                dish_data = re.search("(.*)\((.*)\)", each)
                dish_name = dish_data.group(1)
                dish_price = dish_data.group(2)
                sheet.cell(row=j, column=15+m).value = dish_name
                sheet.cell(row=j, column=15+m+1).value = dish_price
                m += 2

        tag_list = data[i][17].split(' ')
        n = 0
        for each in tag_list:
            if each and n < 5:
                sheet.cell(row=j, column=25+n).value = each

        sheet.cell(row=j, column=30).value = data[i][18]
        sheet.cell(row=j, column=31).value = data[i][19]
        sheet.cell(row=j, column=32).value = data[i][20]
        sheet.cell(row=j, column=33).value = data[i][21]
        sheet.cell(row=j, column=34).value = data[i][22]
        sheet.cell(row=j, column=35).value = data[i][23]
        sheet.cell(row=j, column=36).value = data[i][24]

    excel.save('../../data/shenzhen_dianping.xlsx')


if __name__ == '__main__':
    deal_dashes()
