import json, re
from urllib import request as rq, parse
import os
from flask import Blueprint, current_app
from flask import render_template, make_response, send_from_directory
from flask import request, g
from openpyxl import Workbook
from config import MDianPingConfig

from model.models import MDPModels

dianping = Blueprint('dianping', __name__)


@dianping.before_request
def before_request():
    if not hasattr(g, 'db'):
        user = MDianPingConfig.USER
        password = MDianPingConfig.PASSWORD
        databse = MDianPingConfig.DATABASE
        g.db = MDPModels(databse, user, password)


@dianping.teardown_request
def teardown_request(exception):
    if hasattr(g, 'db'):
        g.db.close_db()


@dianping.route('/dianping')
def index():
    return render_template("dianping.html")


@dianping.route('/dianping/create_table')
def create_table():
    table_name = request.args.get('table_name')
    data = g.db.create_table(table_name)
    return json.dumps(data)


@dianping.route('/dianping/get_params', methods=['POST', 'GET'])
def get_params():
    city_id = request.form['cityid']
    city = request.form['city']
    province = request.form['province']
    cat_id = request.form['catid']
    table_name = request.form['table_name']

    url = "http://localhost:6800/schedule.json"
    data = {
        'project': 'mdianping',
        'spider': 'food',
        'city': city.encode('unicode-escape'),
        'province': province.encode('unicode-escape'),
        'city_id': str(city_id),
        'cat_id': str(cat_id),
        'table_name': table_name
    }
    encode_data = parse.urlencode(data)
    req = rq.Request(url=url, data=encode_data.encode('utf-8'))
    req.add_header("Content-Type", "application/x-www-form-urlencoded;charset=utf-8")
    response = rq.urlopen(req)
    req_data = response.read().decode('utf-8')
    return req_data


@dianping.route('/dianping/get_table')
def get_table():
    conn = g.db.conn
    cur = conn.cursor()
    sql = "select table_name from information_schema.tables"
    cur.execute(sql)
    table_names = cur.fetchall()
    table = []
    for each in table_names:
        if "dp_" in each[0]:
            table.append(each[0])
    return json.dumps(table)


@dianping.route('/dianping/export_table')
def export_table():
    ILLEGAL_CHARACTERS_RE = re.compile('[\000-\010]|[\013-\014]|[\016-\037]')
    table_name = request.args.get('table_name')
    if os.path.isfile(current_app.root_path + "/xlsx/%s.xlsx" % table_name):
        pass
    else:
        conn = g.db.conn
        cur = conn.cursor()
        sql = "select * from %s" % table_name
        cur.execute(sql)
        data = cur.fetchall()
        wbk = Workbook()
        sheet = wbk.create_sheet(table_name, 0)
        sheet.cell(row=1, column=1).value = 'uid'
        sheet.cell(row=1, column=2).value = 'name'
        sheet.cell(row=1, column=3).value = 'branch_name'
        sheet.cell(row=1, column=4).value = 'stars'
        sheet.cell(row=1, column=5).value = 'address'
        sheet.cell(row=1, column=6).value = 'center'
        sheet.cell(row=1, column=7).value = 'price'
        sheet.cell(row=1, column=8).value = 'tel'
        sheet.cell(row=1, column=9).value = 'comments'
        sheet.cell(row=1, column=10).value = 'category_name'
        sheet.cell(row=1, column=11).value = 'sub_category_name'
        sheet.cell(row=1, column=12).value = 'district'
        sheet.cell(row=1, column=13).value = 'region_name'
        sheet.cell(row=1, column=14).value = 'taste'
        sheet.cell(row=1, column=15).value = 'environment'
        sheet.cell(row=1, column=16).value = 'services'
        sheet.cell(row=1, column=17).value = 'has_deals'
        sheet.cell(row=1, column=18).value = 'bookable'
        sheet.cell(row=1, column=19).value = 'has_takeaway'
        sheet.cell(row=1, column=20).value = 'has_mobilepay'
        sheet.cell(row=1, column=21).value = 'has_promote'
        sheet.cell(row=1, column=22).value = 'opentime'
        sheet.cell(row=1, column=23).value = 'city'
        sheet.cell(row=1, column=24).value = 'pro'
        sheet.cell(row=1, column=25).value = 'dish1'
        sheet.cell(row=1, column=26).value = 'dish2'
        sheet.cell(row=1, column=27).value = 'dish3'
        sheet.cell(row=1, column=28).value = 'dish4'
        sheet.cell(row=1, column=29).value = 'dish5'

        i = 2
        for each in data:
            for m in range(1, 23):
                tmp = ILLEGAL_CHARACTERS_RE.sub('', each[m])
                sheet.cell(row=i, column=m).value = tmp
            sheet.cell(row=i, column=23).value = each[24]
            sheet.cell(row=i, column=24).value = each[25]
            dishes = each[23]
            dish_list = dishes.split(' ')
            dish_len = len(dish_list) if len(dish_list) <= 5 else 5
            for x in range(0, dish_len):
                sheet.cell(row=i, column=25+x).value = dish_list[x]
            i += 1

        wbk.save(current_app.root_path + "/xlsx/%s.xlsx" % table_name)
    response = make_response(send_from_directory(current_app.root_path + "/xlsx", table_name+".xlsx"))
    response.headers["Content-Disposition"] = "attachment; filename=%s.xlsx;" % table_name
    return response

