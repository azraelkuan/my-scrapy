import json
from urllib import request as rq, parse
import os, re
from flask import Blueprint, current_app
from flask import render_template, make_response, send_from_directory
from flask import request, g
from openpyxl import Workbook
from config import GaodeConfig

from model.models import Models

gaode = Blueprint('gaode', __name__)


@gaode.before_request
def before_request():
    if not hasattr(g, 'db'):
        user = GaodeConfig.USER
        password = GaodeConfig.PASSWORD
        databse = GaodeConfig.DATABASE
        g.db = Models(databse, user, password)


@gaode.teardown_request
def teardown_request(exception):
    if hasattr(g, 'db'):
        g.db.close_db()


@gaode.route('/')
def main():
    return render_template('main.html')


@gaode.route('/gaode')
def index():
    return render_template("gaode.html")


@gaode.route('/gaode/create_table')
def create_table():
    table_name = request.args.get('table_name')
    data = g.db.create_table(table_name)
    return json.dumps(data)


@gaode.route('/gaode/get_params', methods=['POST', 'GET'])
def get_params():
    try:
        tag = request.form['tag']
    except:
        tag = ""
    try:
        city = request.form['city']
    except:
        city = ""
    province = request.form['province']
    try:
        keywords = request.form['keyword']
    except:
        keywords = ""
    table_name = request.form['table_name']

    url = "http://104.156.239.27:6800/schedule.json"
    data = {
        'project': 'gaode',
        'spider': 'gaode',
        'city_list': city.encode('unicode-escape'),
        'tags': tag,
        'province': province.encode('unicode-escape'),
        'keywords': keywords.encode('unicode-escape'),
        'table_name': table_name
    }
    encode_data = parse.urlencode(data)
    req = rq.Request(url=url, data=encode_data.encode('utf-8'))
    req.add_header("Content-Type", "application/x-www-form-urlencoded;charset=utf-8")
    response = rq.urlopen(req)
    req_data = response.read().decode('utf-8')

    return req_data


@gaode.route('/gaode/get_table')
def get_table():
    conn = g.db.conn
    cur = conn.cursor()
    sql = "select table_name from information_schema.tables"
    cur.execute(sql)
    table_names = cur.fetchall()
    table = []
    for each in table_names:
        if "web_" in each[0]:
            table.append(each[0])
    return json.dumps(table)


@gaode.route('/gaode/export_table')
def export_table():
    ILLEGAL_CHARACTERS_RE = re.compile('[\000-\010]|[\013-\014]|[\016-\037]')
    table_name = request.args.get('table_name')
    if os.path.isfile(current_app.root_path + "/xlsx/%s.xlsx" % table_name):
        pass
    else:
        conn = g.db.conn
        cur = conn.cursor()
        sql = "select * from %s" % table_name
        cur.execute(sql)
        data = cur.fetchall()
        wbk = Workbook()
        sheet = wbk.create_sheet(table_name, 0)
        sheet.cell(row=1, column=1).value = 'uid'
        sheet.cell(row=1, column=2).value = 'name'
        sheet.cell(row=1, column=3).value = 'address'
        sheet.cell(row=1, column=4).value = 'tag'
        sheet.cell(row=1, column=5).value = 'small tag'
        sheet.cell(row=1, column=6).value = 'location'
        sheet.cell(row=1, column=7).value = 'tel'
        sheet.cell(row=1, column=8).value = 'pro_name'
        sheet.cell(row=1, column=9).value = 'pro_center'
        sheet.cell(row=1, column=10).value = 'city_name'
        sheet.cell(row=1, column=11).value = 'city_center'
        sheet.cell(row=1, column=12).value = 'district_name'
        sheet.cell(row=1, column=13).value = 'district_center'
        sheet.cell(row=1, column=14).value = 'photo_exists'
        sheet.cell(row=1, column=15).value = 'photo1'
        sheet.cell(row=1, column=16).value = 'photo2'
        sheet.cell(row=1, column=17).value = 'photo3'

        i = 2
        for each in data:
            for m in range(1, 18):
                tmp = ILLEGAL_CHARACTERS_RE.sub('', each[m])
                sheet.cell(row=i, column=m).value = tmp
            i += 1
        wbk.save(current_app.root_path + "/xlsx/%s.xlsx" % table_name)
    response = make_response(send_from_directory(current_app.root_path + "/xlsx", table_name+".xlsx"))
    response.headers["Content-Disposition"] = "attachment; filename=%s.xlsx;" % table_name
    return response

    # if os.path.isfile(os.path.join('./xlsx', table_name+".xlsx")):
    #     return send_from_directory('./xlsx', table_name+".xlsx", as_attachment=True)
    # else:
    #     return "fail"

